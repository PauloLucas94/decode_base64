"""
╔══════════════════════════════════════════════════════════════╗
║        CONVERSOR DE FOTOS → BASE64 — SENAI-SP               ║
╠══════════════════════════════════════════════════════════════╣
║  COMO USAR:                                                  ║
║  1. Coloque este script na mesma pasta das fotos             ║
║  2. Crie uma subpasta chamada "fotos"                        ║
║  3. Nomeie as fotos como:                                    ║
║       auditorio-101-1.jpg  auditorio-101-2.png  etc.        ║
║  4. Execute:  python converter_simples.py                    ║
║  5. Abra o arquivo resultado.xlsx gerado                     ║
║  6. Copie e cole os valores na sua planilha                  ║
╚══════════════════════════════════════════════════════════════╝
"""

import os
import base64
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PASTA_FOTOS = "fotos"
EXTENSOES   = [".jpg", ".jpeg", ".png", ".webp"]

# ── Helpers ────────────────────────────────────────────────────────────────────
def para_base64(caminho):
    ext  = Path(caminho).suffix.lower()
    mime = {".jpg": "image/jpeg", ".jpeg": "image/jpeg",
            ".png": "image/png",  ".webp": "image/webp"}.get(ext, "image/jpeg")
    with open(caminho, "rb") as f:
        dados = base64.b64encode(f.read()).decode("utf-8")
    return f"data:{mime};base64,{dados}"

def encontrar_foto(pasta, cfp, numero):
    for ext in EXTENSOES:
        c = os.path.join(pasta, f"auditorio-{cfp}-{numero}{ext}")
        if os.path.exists(c):
            return c
    return None

# ── Verificação ────────────────────────────────────────────────────────────────
if not os.path.exists(PASTA_FOTOS):
    print(f"❌  Pasta '{PASTA_FOTOS}' não encontrada.")
    exit(1)

# ── Descobrir unidades pela pasta ─────────────────────────────────────────────
cfps = set()
for arq in os.listdir(PASTA_FOTOS):
    nome = Path(arq).stem  # ex: auditorio-101-1
    partes = nome.split("-")
    if len(partes) == 3 and partes[0] == "auditorio":
        cfps.add(partes[1])  # ex: 101

cfps = sorted(cfps)
if not cfps:
    print("❌  Nenhuma foto encontrada no padrão 'auditorio-{CFP}-{N}.ext'")
    exit(1)

print(f"\n📋  Unidades encontradas: {', '.join(cfps)}\n")

# ── Criar Excel de resultado ───────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Base64"

# Estilos
HDR_BG  = "1A1A1A"
HDR_FG  = "FFFFFF"
ALT     = "F5F5F5"
thin    = Side(style="thin", color="CCCCCC")
borda   = Border(left=thin, right=thin, top=thin, bottom=thin)

def estilo_header(cell):
    cell.font      = Font(name="Arial", bold=True, color=HDR_FG, size=10)
    cell.fill      = PatternFill("solid", start_color=HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = borda

def estilo_dado(cell, alt=False):
    cell.font      = Font(name="Arial", size=9)
    cell.fill      = PatternFill("solid", start_color=ALT if alt else "FFFFFF")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = borda

# Cabeçalhos
cabecalhos = ["Cod_Unidade", "Foto_01_B64", "Foto_02_B64", "Foto_03_B64", "Foto_04_B64"]
larguras   = [14, 30, 30, 30, 30]

ws.row_dimensions[1].height = 22
for i, (cab, larg) in enumerate(zip(cabecalhos, larguras), start=1):
    c = ws.cell(row=1, column=i, value=cab)
    estilo_header(c)
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = larg

ws.freeze_panes = "B2"

# ── Preencher linhas ───────────────────────────────────────────────────────────
ok_total  = 0
nok_total = 0

for r, cfp in enumerate(cfps, start=2):
    alt = (r % 2 == 0)
    ws.row_dimensions[r].height = 16

    # Cod_Unidade — reconstrói o formato com ponto (ex: 101 → 1.01)
    cod_fmt = cfp if "." in cfp else (cfp[:-2] + "." + cfp[-2:]) if len(cfp) > 2 else cfp
    c = ws.cell(row=r, column=1, value=cod_fmt)
    estilo_dado(c, alt)

    print(f"   📋  Unidade {cod_fmt}")

    for foto_num in range(1, 5):
        caminho = encontrar_foto(PASTA_FOTOS, cfp, foto_num)
        col = foto_num + 1  # colunas B, C, D, E

        if caminho:
            b64 = para_base64(caminho)
            cell = ws.cell(row=r, column=col, value=b64)
            estilo_dado(cell, alt)
            kb = os.path.getsize(caminho) / 1024
            print(f"      ✅  Foto {foto_num}: {Path(caminho).name} ({kb:.0f} KB)")
            ok_total += 1
        else:
            cell = ws.cell(row=r, column=col, value="")
            estilo_dado(cell, alt)
            print(f"      ⚠️   Foto {foto_num}: não encontrada")
            nok_total += 1

# ── Salvar ─────────────────────────────────────────────────────────────────────
wb.save("resultado_base64.xlsx")

print(f"""
╔══════════════════════════════════════════════════════════════╗
║  ✅  CONCLUÍDO!                                              ║
╠══════════════════════════════════════════════════════════════╣
║  Fotos convertidas : {str(ok_total).ljust(39)} ║
║  Não encontradas   : {str(nok_total).ljust(39)} ║
║  Arquivo gerado    : resultado_base64.xlsx                   ║
╠══════════════════════════════════════════════════════════════╣
║  PRÓXIMO PASSO:                                              ║
║  1. Abra resultado_base64.xlsx                               ║
║  2. Copie as colunas Foto_01_B64 ... Foto_04_B64             ║
║  3. Cole na sua planilha principal                           ║
╚══════════════════════════════════════════════════════════════╝
""")